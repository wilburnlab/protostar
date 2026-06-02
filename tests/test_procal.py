"""Tests for the PROCAL calibrant reference.

``load_procal`` / ``procal_sequences`` read the committed parquet (always
available). ``parse_procal_supplement`` is exercised against a tiny synthetic
xlsx that mimics the ``PROCAL Sequences`` sheet layout (the real Table S1 ships a
malformed defined-name spec that breaks openpyxl, which is exactly why the parser
reads the OOXML zip directly rather than via openpyxl).
"""

from __future__ import annotations

import openpyxl
import pytest

from protostar.peptides import procal
from protostar.peptides.procal import PROCAL_TABLE

# ── committed parquet ───────────────────────────────────────────────────


def test_load_procal_has_40_rows_and_schema():
    t = procal.load_procal()
    assert t.num_rows == 40
    assert t.schema.names == PROCAL_TABLE.names


def test_procal_sequences_known_members():
    seqs = procal.procal_sequences()
    assert len(seqs) == 40
    assert isinstance(seqs, frozenset)
    assert "YSAHEEHHYDK" in seqs  # peptide #1
    assert "SLIFFLSTLLK" in seqs  # peptide #40


def test_procal_mz_present():
    t = procal.load_procal()
    row0 = t.slice(0, 1).to_pylist()[0]
    assert row0["sequence"] == "YSAHEEHHYDK"
    assert abs(row0["mz_2plus"] - 708.3024) < 1e-3


# ── parser (synthetic fixture) ──────────────────────────────────────────


def _write_procal_like(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="PROCAL Sequences")
    ws.append(["PROCAL peptide sequences and properties."])  # row 1 title
    ws.append([None])  # row 2 spacer
    ws.append(  # row 3 header (mixed-case "Sequence" → skipped by the parser)
        [
            "Peptide Number",
            "Sequence",
            "(M+2H)2+",
            "(M+3H)3+",
            "Expected Charge",
            "SSRCalc HI",
            "SSRCalc HI alt",
            "RT 60 min",
        ]
    )
    ws.append([1, "YSAHEEHHYDK", 708.3024, 472.5373, "2;3", -1.29, 1.45, "9.96+/-0.17"])
    ws.append([2, "HEHISSDYAGK", 622.2887, 415.1949, "2;3", 0.89, 5.92, "13.02+/-0.14"])
    wb.save(path)


def test_parse_procal_supplement(tmp_path):
    xlsx = tmp_path / "procal.xlsx"
    _write_procal_like(xlsx)
    t = procal.parse_procal_supplement(xlsx)
    assert t.num_rows == 2
    assert t.schema.names == PROCAL_TABLE.names
    rows = t.to_pylist()
    assert [r["sequence"] for r in rows] == ["YSAHEEHHYDK", "HEHISSDYAGK"]
    assert rows[0]["peptide_number"] == 1
    assert abs(rows[0]["mz_2plus"] - 708.3024) < 1e-6
    assert rows[0]["expected_charge"] == "2;3"
    assert abs(rows[1]["ssrcalc_hi"] - 0.89) < 1e-6


def test_parse_procal_missing_sheet_raises(tmp_path):
    xlsx = tmp_path / "wrong.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Not PROCAL"
    wb.save(xlsx)
    with pytest.raises(ValueError, match="not found"):
        procal.parse_procal_supplement(xlsx)
