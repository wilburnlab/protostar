"""Parse the ProteomeTools manuscript supplements into the peptide reference.

The three source papers publish their synthetic-peptide lists as supplementary
Excel workbooks. These are a **one-time manual input** — download them from the
articles (redistribution is not ours to make) and point this parser at them:

    Zolg2017     41592_2017_BFnmeth4153_MOESM249_ESM.xlsx  (Nat. Methods 2017)
    Gessulat2019 41592_2019_426_MOESM3_ESM.xlsx            (Nat. Methods 2019)
    Wilhelm2021  41467_2021_23713_MOESM4_ESM.xlsx          (Nat. Commun. 2021)

Each workbook has one sheet per peptide *set*, every sheet carrying a
``Pool name`` + ``Sequence`` column (the join we need). This module maps each
dataset's sheets to a canonical ``set`` label and flattens them into the
``PEPTIDE_REFERENCE_TABLE`` (see ``reference.py``); ``pipelines/05_peptide_reference.py``
drives it and writes the committed parquet.

Parsing needs ``openpyxl`` (the ``peptides`` optional extra) — a build-time-only
dependency; reading the committed parquet does not.

Provenance / decisions:
- Bare ``sequence`` as published; no charge, no modseq, no m/z (deferred to use).
- ``set`` tags the originating sheet (proteotypic, tmt_proteotypic, hla_class_i, …).
- Zolg's "Quality Control Peptides" sheet is the calibrant set (JPT-RT / JPT-QC /
  Pierce-RT); flagged ``is_qc`` with ``qc_type``. NOTE: the PROCAL-40 subset is
  *not* separately labeled in this supplement (it is 40 of the JPT calibrants);
  pinning the exact 40 needs the PROCAL paper (PXD006832) and is a follow-up.
- RT/iRT carried only for Wilhelm (its ``Identifications`` sheet publishes
  ``Average Retention Time`` + ``Average iRT``), joined by sequence.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import pyarrow as pa

#: Default Excel filename per dataset (the published supplement). Path is supplied
#: at call time (the workbooks are a manual, un-redistributed input).
SUPPLEMENT_FILES: dict[str, str] = {
    "Zolg2017": "41592_2017_BFnmeth4153_MOESM249_ESM.xlsx",
    "Gessulat2019": "41592_2019_426_MOESM3_ESM.xlsx",
    "Wilhelm2021": "41467_2021_23713_MOESM4_ESM.xlsx",
}


@dataclass(frozen=True, slots=True)
class SheetSpec:
    """One peptide sheet → canonical ``set`` label."""

    title: str  # exact worksheet title (Zolg's are quoted literally)
    set_label: str
    is_qc: bool = False


@dataclass(frozen=True, slots=True)
class DatasetSpec:
    """How to parse one dataset's workbook."""

    dataset: str
    sheets: tuple[SheetSpec, ...]
    # Optional sequence→RT/iRT sheet (Wilhelm only).
    rt_sheet: str | None = None
    rt_cols: tuple[str, str] = field(default=("average retention time", "average irt"))


DATASET_SPECS: dict[str, DatasetSpec] = {
    "Zolg2017": DatasetSpec(
        "Zolg2017",
        (
            SheetSpec('"Proteotypic Set"', "proteotypic"),
            SheetSpec('"Missing Gene Set"', "missing_gene"),
            SheetSpec('"SRMAtlas Set"', "srmatlas"),
            SheetSpec("Quality Control Peptides", "qc", is_qc=True),
        ),
    ),
    "Gessulat2019": DatasetSpec(
        "Gessulat2019",
        (
            SheetSpec("Isoform", "isoform"),
            SheetSpec("TMT Proteotypic", "tmt_proteotypic"),
            SheetSpec("Addon Missing Genes", "addon_missing"),
            SheetSpec("Missing in first", "missing_first"),
        ),
    ),
    "Wilhelm2021": DatasetSpec(
        "Wilhelm2021",
        (
            SheetSpec("HLA Class I", "hla_class_i"),
            SheetSpec("HLA Class II", "hla_class_ii"),
            SheetSpec("AspN", "aspn"),
            SheetSpec("LysN", "lysn"),
        ),
        rt_sheet="Identifications",
    ),
}


def _norm_header(header: tuple) -> dict[str, int]:
    """Normalized (lowercased, de-quoted, trimmed) header → column index."""
    out: dict[str, int] = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        key = str(h).strip().strip('"').lower()
        if key and key not in out:
            out[key] = i
    return out


def _pool_prefix(pool: str | None) -> str | None:
    from .reference import pool_prefix

    return pool_prefix(pool) if pool else None


def _load_rt_map(wb, spec: DatasetSpec) -> dict[str, tuple[float | None, float | None]]:
    """sequence → (rt, irt) from the dataset's RT sheet, if any."""
    if spec.rt_sheet is None or spec.rt_sheet not in wb.sheetnames:
        return {}
    ws = wb[spec.rt_sheet]
    it = ws.iter_rows(values_only=True)
    cm = _norm_header(next(it))
    seq_i = cm.get("sequence")
    rt_i = cm.get(spec.rt_cols[0])
    irt_i = cm.get(spec.rt_cols[1])
    if seq_i is None:
        return {}

    def _f(row, i):
        if i is None or i >= len(row) or row[i] is None:
            return None
        try:
            v = float(row[i])
        except (TypeError, ValueError):
            return None
        return None if v != v else v  # drop NaN

    out: dict[str, tuple[float | None, float | None]] = {}
    for row in it:
        if seq_i >= len(row) or not row[seq_i]:
            continue
        out[str(row[seq_i]).strip()] = (_f(row, rt_i), _f(row, irt_i))
    return out


def parse_supplement(xlsx_path: "str | Path", dataset: str) -> pa.Table:
    """Parse one dataset's Excel supplement into a ``PEPTIDE_REFERENCE_TABLE``.

    Requires ``openpyxl`` (the ``peptides`` extra). ``peptide_id`` is a stable
    0-based row index over the sheets in spec order.
    """
    import openpyxl  # build-time-only dep (peptides extra)

    from .reference import PEPTIDE_REFERENCE_TABLE

    spec = DATASET_SPECS[dataset]
    wb = openpyxl.load_workbook(Path(xlsx_path), read_only=True, data_only=True)
    rt_map = _load_rt_map(wb, spec)

    cols: dict[str, list] = {name: [] for name in PEPTIDE_REFERENCE_TABLE.names}
    pid = 0
    for sheet in spec.sheets:
        if sheet.title not in wb.sheetnames:
            raise ValueError(
                f"{dataset}: sheet {sheet.title!r} not in workbook (have {wb.sheetnames})"
            )
        ws = wb[sheet.title]
        it = ws.iter_rows(values_only=True)
        cm = _norm_header(next(it))
        seq_i = cm.get("sequence")
        pool_i = cm.get("pool name")  # absent on the QC sheet
        qctype_i = cm.get("qc type")
        if seq_i is None:
            raise ValueError(f"{dataset}/{sheet.title}: no 'Sequence' column")
        for row in it:
            if seq_i >= len(row) or not row[seq_i] or not str(row[seq_i]).strip():
                continue
            seq = str(row[seq_i]).strip()
            pool = (
                str(row[pool_i]).strip()
                if pool_i is not None and pool_i < len(row) and row[pool_i]
                else None
            )
            qc_type = (
                str(row[qctype_i]).strip()
                if sheet.is_qc and qctype_i is not None and qctype_i < len(row) and row[qctype_i]
                else None
            )
            rt, irt = rt_map.get(seq, (None, None))
            cols["peptide_id"].append(pid)
            cols["sequence"].append(seq)
            cols["pool"].append(pool)
            cols["pool_prefix"].append(_pool_prefix(pool))
            cols["set"].append(sheet.set_label)
            cols["is_qc"].append(sheet.is_qc)
            cols["qc_type"].append(qc_type)
            cols["rt"].append(rt)
            cols["irt"].append(irt)
            pid += 1
    wb.close()
    return pa.table(
        {k: pa.array(v, type=PEPTIDE_REFERENCE_TABLE.field(k).type) for k, v in cols.items()},
        schema=PEPTIDE_REFERENCE_TABLE,
    )


__all__ = [
    "DATASET_SPECS",
    "DatasetSpec",
    "SUPPLEMENT_FILES",
    "SheetSpec",
    "parse_supplement",
]
